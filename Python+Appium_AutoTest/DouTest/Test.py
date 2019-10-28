import time
from appium import webdriver
import random
import openpyxl


wb = openpyxl.load_workbook(
    "C:\\Users\\admin\\Desktop\\PythonCode\\Python+Appium_AutoTest\\DouTest\\AppiumTestNeed.xlsx")
ws = wb.active


def dc_option():
    desire_caps = {}
    desire_caps['platformName'] = ws["A2"].value
    desire_caps['playformVersion'] = ws["B2"].value
    desire_caps['deviceName'] = ws["C2"].value
    desire_caps['automationName'] = ws["D2"].value
    desire_caps['appPackage'] = ws["E2"].value
    desire_caps['appActivity'] = ws["F2"].value
    desire_caps['appWaitActivity'] = ws["G2"].value
    desire_caps['newCommandTimeout'] = ws["H2"].value
    desire_caps['resetKeyboard'] = ws["I2"].value
    return desire_caps


WifiName = ["EM-public", "ASUS"]
WifiList = ["three bars", "signal full"]
size = {}
driver = webdriver.Remote("http://localhost:4723/wd/hub", dc_option())
# --------------------------------------Before Test-----------------------------------------------


class Tools(object):
    def get_size(self):
        x = driver.get_window_size()['width']
        y = driver.get_window_size()['height']
        print(x, y)
        return(x, y)

    def Swipe_Menu(self, size, time):  # 上拉菜单
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.2)
        x_2 = int(size[0]*0.3)
        y_1 = int(size[1]*0.75)
        y_2 = int(size[1]*0.15)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def Swipe_down(self, size, time):  # 横屏下拉状态栏
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.5)
        x_2 = int(size[0]*0.55)
        y_1 = int(size[1]*0.01)
        y_2 = int(size[1]*0.4)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def Swipe_down(self, size, time):  # 竖屏时的下拉状态栏
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.9)
        x_2 = int(size[0]*0.8)
        y_1 = int(size[1]*0.01)
        y_2 = int(size[1]*0.4)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def Swpie_Up(self, size, time):  # 横屏时的上拉状态栏
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.2)
        x_2 = int(size[0]*0.3)
        y_1 = int(size[1]*0.75)
        y_2 = int(size[1]*0.15)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def Swipe_Up_1(self, size, time):  # 竖屏时的上拉状态栏
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.9)
        x_2 = int(size[0]*0.8)
        y_1 = int(size[1]*0.4)
        y_2 = int(size[1]*0.01)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def Swipe_Right(self, size, time):  # 解锁网易云音乐的锁屏界面
        self.size = size
        self.time = time
        x_1 = int(size[0]*0.1)
        x_2 = int(size[0]*0.8)
        y_1 = int(size[1]*0.35)
        y_2 = int(size[1]*0.36)
        driver.swipe(x_1, y_1, x_2, y_2, time)

    def OpenGetPower(self):
        self.size = size
        driver.find_element_by_accessibility_id("GetPower").click()
        driver.find_element_by_id('com.example.GetPower:id/getPower').click()
        driver.press_keycode(3)

    def CloseWifi(self):
        driver.find_element_by_xpath('//android.widget.Switch'
                                     + '[@content-desc="Wi-Fi,Wi-Fi '
                                     + WifiList[1]
                                     + '.,'
                                     + WifiName[1]
                                     + '"]'
                                     + '/android.widget.FrameLayout/android.view.ViewGroup/'
                                     + 'android.widget.FrameLayout/android.widget.ImageView').click()

    def OpenWifi(self):
        size = Tools().get_size()
        Tools().Swipe_down(size, 1000)
        time.sleep(2)
        driver.find_element_by_xpath('//android.widget.Switch[@content-desc="Wi-Fi,"]/'
                                     + 'android.widget.FrameLayout/android.view.ViewGroup').click()

    def CheckGetPower(self, size):
        self.size = size
        Tools().swipe_top(size, 1000)
        AppStatus = driver.is_app_installed("com.example.GetPower")
        if AppStatus:
            return
        else:
            driver.install_app(
                "https://raw.githubusercontent.com/flashnames/ApkRepository/master/GetPower.apk?token=ALIUASK3HON674SDBP42F5K5LO6PO")
# ------------------------------Class Tools End------------------------------------------------------------

# -------------------------Class Test Start---------------------------------------------------


class Test(object):

    # --------------------------Tencent Video start------------------
    def TententVideo(self, size):
        self.size = size
        time.sleep(4)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(3)
        driver.find_element_by_accessibility_id('Tencent Video').click()
        time.sleep(15)
        driver.find_element_by_id('com.tencent.qqlive:id/a76').click()
        time.sleep(4)
        Video = driver.find_elements_by_id('com.tencent.qqlive:id/cg')
        time.sleep(3)
        driver.find_element_by_xpath(
            '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.support.v7.widget.RecyclerView/android.widget.LinearLayout[2]/android.widget.GridView/android.widget.RelativeLayout[1]/android.widget.TextView').click()
        time.sleep(110)
        driver.find_element_by_id('com.tencent.qqlive:id/bkb').click()
        time.sleep(5)
        driver.find_element_by_id('com.tencent.qqlive:id/bts').click()
        time.sleep(2)
        driver.press_keycode(4)
        time.sleep(2)
        driver.press_keycode(4)
        time.sleep(2)
        driver.press_keycode(4)

    def OpenTencentVideo(self, size):
        self.size = size
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        try:
            Tools().OpenWifi()  # 打开wifi
        except:
            pass
        Tools().Swpie_Up(size, 1000)
        time.sleep(4)
        Tools().Swipe_Menu(size, 1000)
        Test().TententVideo(size)
        size = Tools().get_size()
        Tools().Swipe_down(size, 1000)  # 下拉状态框
        Tools().CloseWifi()  # 关闭wifi
        Tools().Swipe_Up_1(size, 1000)  # 上拉状态框
        driver.press_keycode(3)  # 调用home
        time.sleep(2)
        size = Tools().get_size()
        Tools().Swipe_Menu(size, 1000)  # 上拉屏幕
        time.sleep(2)
        Tools().OpenGetPower()  # 获取电量
        print("Tencent Video Test Finish")
# ---------------------Tencent Video End-------------------------------------------------------------
# ---------------------Open Game start-----------------------------

    def OpenGame(self, size):
        self.size = size
        Tools().Swipe_Menu(size, 1000)
        driver.find_element_by_accessibility_id('NFS Most Wanted').click()
        time.sleep(10)
        driver.press_keycode(3)
        print("Game Test Finish")
# ---------------------Open Game End-----------------------------
# ---------------------OpenCamera start-------------------------

    def record(self):
        self.size = size
        time.sleep(4)
        driver.find_element_by_accessibility_id('Video').click()
        driver.find_element_by_accessibility_id('Video').click()
        time.sleep(5)
        driver.find_element_by_id(
            'com.mediatek.camera:id/video_stop_shutter').click()
        time.sleep(5)
        driver.press_keycode(3)
        return

    def OpenCamera(self, size):
        self.size = size
        Tools().Swipe_Menu(size, 1000)
        driver.find_element_by_accessibility_id('Camera').click()
        for i in range(0, 3):
            time.sleep(3)
            driver.press_keycode(27)
            i = i+1
        Test().record()
        time.sleep(2)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(2)
        Tools().OpenGetPower()
        print("Camera Test Finish")
# ---------------------------OpenCamera end--------------------------
# ---------------------------OpenMusic start---------------------------- --

    def OpenNetEase(self, size):
        self.size = size
        Tools().Swipe_Menu(size, 1000)
        driver.find_element_by_accessibility_id('NetEase Music').click()
        time.sleep(3)
        driver.find_element_by_accessibility_id('我的音乐').click()
        time.sleep(3)
        driver.find_element_by_xpath(
            '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/android.widget.FrameLayout/androidx.drawerlayout.widget.DrawerLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/androidx.viewpager.widget.ViewPager/android.view.ViewGroup/android.widget.FrameLayout/android.widget.ListView/android.widget.RelativeLayout[1]').click()
        time.sleep(3)
        driver.find_element_by_id('com.netease.cloudmusic:id/bed').click()
        driver.press_keycode(26)
        time.sleep(3)
        driver.press_keycode(26)
        time.sleep(3)
        driver.find_element_by_id('com.netease.cloudmusic:id/as6').click()
        Tools().Swipe_Right(size, 2000)
        time.sleep(3)
        driver.press_keycode(3)
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        Tools().OpenGetPower()

 #  def OpenGoogleMusic(self,size):
        self.size = size
        Tools().Swipe_Menu(size, 1000)
        driver.find_element_by_accessibility_id('Play Music').click()
        driver.find_element_by_id('Show navigation drawer').click()
        time.sleep(3)
        driver.find_element_by_xpath(
            '/hierarchy/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.widget.LinearLayout/android.widget.FrameLayout/android.support.v4.widget.DrawerLayout/android.widget.LinearLayout/android.widget.ListView/android.widget.TextView[2]').click()
        driver.find_element_by_id(
            'com.google.android.music:id/li_play_button').click()
        driver.press_keycode(26)
        time.sleep(3)
        driver.press_keycode(26)
        time.sleep(4)
        Tools().Swipe_Menu(size, 1000)
        Choose = driver.find_element_by_id(
            'com.google.android.music:id/li_play_button').text
        if Choose == "Pause List Added":
            driver.find_element_by_id(
                'com.google.android.music:id/li_play_button').click()
        time.sleep(3)
        driver.press_keycode(3)
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        Tools().OpenGetPower()
        print("Music Test Finish")
# ---------------------------OpenMusic End------------------------------
# ---------------------------OpenBrowser start-----------------------------

    def OpenBrowser(self, size):
        self.size = size
        time.sleep(3)
        try:
            Tools().OpenWifi()
        except:
            pass
        time.sleep(3)
        Tools().Swpie_Up(size, 1000)
        size = Tools().get_size()
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(3)
        driver.find_element_by_accessibility_id('Chrome').click()
        driver.find_element_by_id('com.android.browser:id/url').clear()
        driver.find_element_by_id(
            'com.android.browser:id/url').send_keys('http://www.tclcom.com/test/')
        driver.press_keycode(66)
        time.sleep(10)
        driver.press_keycode(3)
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(3)
        Tools().OpenGetPower()
        print("Browser Test Finish")
# --------------------------OpenBrowser end--------------------------------

    def OpenWeiBo(self, size):
        self.size = size
        time.sleep(3)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(3)
        driver.find_element_by_accessibility_id('Weibo').click()
        time.sleep(3)
        size = Tools().get_size()
        for second in range(10):
            time.sleep(3)
            Tools().Swipe_Menu(size, 1000)
            second = second+1
            print(second)
        time.sleep(3)
        driver.press_keycode(3)
        size = Tools().get_size()
        time.sleep(10)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(3)
        Tools().OpenGetPower()
        print("WeiBo Test Finish")
# -------------------------OpenWeiBo end-------------------------------
# -------------------------OpenWeChat start----------------------------

    def OpenWeChat(self, size):
        self.size = size
        Tools().Swipe_Menu(size, 1000)
        driver.find_element_by_accessibility_id('WeChat').click()
        time.sleep(10)
        driver.find_element_by_id('com.tencent.mm:id/bai').click()
        driver.find_element_by_accessibility_id('Comments').click()
        for i in range(0, 3):
            Rnum = random.randint(0, 29)
            driver.find_element_by_id('com.tencent.mm:id/aqc').clear()
            driver.find_element_by_id(
                'com.tencent.mm:id/aqc').send_keys('Message%d' % Rnum)
            driver.find_element_by_id('com.tencent.mm:id/aqj').click()
            time.sleep(60)
            i = i+1
            print(i)
        driver.press_keycode(3)
        Tools().Swipe_Menu(size, 1000)
        time.sleep(2)
        Tools().OpenGetPower()
        time.sleep(2)
        Tools().CloseWifi()
        print("WeChat Test Finish")
# ----------------------------OpenWeChat end-----------------------------
# ----------------------------All Test Should be in here to Run------------------

    def Test_init(self):
        size = Tools().get_size()
        # Test().OpenGame(size)
        # Test().OpenGoogleMusic(size)
        # Test().OpenNetEase(size)
        # Test().OpenTencentVideo(size)
        # Test().OpenCamera(size)
        # Test().OpenBrowser(size)
        # Test().OpenWeiBo(size)
        Test().OpenWeChat(size)


# -----------------------------------------------------------------------------------------
# --------------------------Main Menu--------------------------------------
if __name__ == "__main__":
    Test()
    Tools()
    time.sleep(3)
    for i in range(13):
        Test().Test_init()
