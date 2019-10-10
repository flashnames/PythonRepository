import openpyxl

class Test(object):#Just Only Use to Test
    def CheckWorkBookTitle(self):#读取工作表中表格标题
        wb=openpyxl.load_workbook("Test.xlsx")
        sheets=wb.sheetnames
        print(sheets,type(sheets))
        sheet_1=wb.get_sheet_by_name("Sheet1")
        for  sheet in wb:
            print(sheet.title)
    def ReadData(self): #读取工作表中单元格数据
         wb=openpyxl.load_workbook("Test.xlsx")
         ws=wb.active
         print(ws["A1"].value)
if __name__=="__main__":
    Test().ReadData()